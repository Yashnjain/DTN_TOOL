from openpyxl import Workbook
from openpyxl.styles import PatternFill,Alignment
from .blob import upload_blob

def set_cell(ws, cell, value, fill, alignment):
    ws[cell] = value
    ws[cell].fill = fill
    ws[cell].alignment = alignment

def create_excel_file(location_prices : list,effective_date : str,path : str,effective_time : str) -> None:
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "BioUrjaMagellanPrice"
        yellow_fill = PatternFill (start_color='00FFFF00', end_color='00FFFF00', fill_type='solid')
        ws["A1"] = "BioUrja Trading, LLC"
        ws["A1"].alignment = Alignment(horizontal='left')
        ws["A2"] = effective_date + " " + effective_time
        ws["A2"].alignment = Alignment(horizontal='left')
        ws["A2"].fill = yellow_fill
        
        temp = 3

        for record in location_prices:
            ws[f"A{temp}"] = record["location"] + " " + record["state"] + " " + "Magellan"
            ws[f"A{temp}"].alignment = Alignment(horizontal='left')
            ws[f"A{temp + 1}"] = "PRICE"
            ws[f"A{temp + 1}"].alignment =  Alignment(horizontal='left')
            ws[f"A{temp + 2}"] = "CHANGE"
            ws[f"A{temp + 2}"].alignment =  Alignment(horizontal='left')
            ws[f"B{temp}"] = "Ethanol"
            ws[f"B{temp}"].alignment =  Alignment(horizontal='left')
            ws[f"B{temp + 1}"] = record["price"]
            ws[f"B{temp + 1}"].alignment =  Alignment(horizontal='left')
            ws[f"B{temp + 2}"] = record["change"]
            ws[f"B{temp + 2}"].alignment = Alignment(horizontal='left')
            ws[f"B{temp + 1}"].fill = yellow_fill
            ws[f"B{temp + 2}"].fill = yellow_fill
            temp += 4
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width        
        # ws[f"A{temp + 1}"] = "<!--END OF FILE-->"   
        # ws[f"A{temp + 1}"].alignment = Alignment(horizontal='left')
        wb.save(path)
        upload_blob(path)
    except Exception as e:
        raise Exception (f"Exception during Excel creation {e}")    
    
    
    